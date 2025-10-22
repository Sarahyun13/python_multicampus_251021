import openpyxl
import random

# 새로운 워크북 생성
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "전자제품목록"

# 헤더 추가
headers = ["제품ID", "제품명", "가격", "수량"]
for col, header in enumerate(headers, 1):
    ws.cell(row=1, column=col, value=header)

# 제품 데이터 생성을 위한 샘플 데이터
product_names = [
    "스마트폰", "노트북", "태블릿", "스마트워치", "무선이어폰", 
    "블루투스 스피커", "공기청정기", "스마트TV", "게이밍모니터", "냉장고"
]

# 100개의 제품 데이터 생성
for row in range(2, 102):  # 2행부터 101행까지
    product_id = f"P{str(row-1).zfill(3)}"  # P001, P002, ...
    product_name = random.choice(product_names)
    price = random.randint(100000, 2000000)  # 10만원 ~ 200만원
    quantity = random.randint(1, 100)  # 1 ~ 100개
    
    # 데이터 입력
    ws.cell(row=row, column=1, value=product_id)
    ws.cell(row=row, column=2, value=product_name)
    ws.cell(row=row, column=3, value=price)
    ws.cell(row=row, column=4, value=quantity)

# 열 너비 자동 조정
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter
    for cell in col:
        if len(str(cell.value)) > max_length:
            max_length = len(str(cell.value))
    ws.column_dimensions[column].width = max_length + 2

# 파일 저장
wb.save("ProductList.xlsx")
print("ProductList.xlsx 파일이 생성되었습니다.")